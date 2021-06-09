#!/usr/bin/env python3

# Project: OSSEM Data Dictionaries
# Author: Jose Rodriguez (@Cyb3rPandaH)
# License: GPLv3

# Importing libraries
import yaml
yaml.Dumper.ignore_aliases = lambda *args : True

import glob
import os
from os import path

import openpyxl

# Creating a list with ms excel files' names in your current directory
excel_files = glob.glob(path.join(path.dirname(__file__),"*.xlsx"))

# Parsing every ms excel file in your current directory
for excel_file_path in excel_files:
    # Getting name of file (includes extensions such as .xlsx)
    file_name = os.path.basename(excel_file_path)
    # Getting name of file (without extensions such as .xlsx)
    file_name = file_name[:file_name.find('.')]
    # Getting content of ms excel file
    wb = openpyxl.load_workbook(excel_file_path)
    sheetnames = wb.sheetnames
    # Parsing every sheet within the file
    for sheet in sheetnames:
        # Defining sheet to parse
        sheet_to_parse = wb[sheet]
        # Defining a control variable to parse sections of the ms excel file
        control = ''
        # Defining references
        event_fields_yaml = []
        references_yaml = []
        tags_yaml = []
        for line in sheet_to_parse.iter_rows(values_only=True):
            # Updating control variable to identify sections of the ms excel file       
            if line[0] == 'standard_name':
                control = 'data dictionary'
                continue
            if line[0] == 'references':
                control = 'references'
                continue
            if line[0] == 'tags':
                control = 'tags'
                continue
            # Getting values to create yaml file
            if line[0] == 'title':
                title_yaml = line[1].rstrip()
            if line[0] == 'description':
                description_yaml = line[1].rstrip()
            if line[0] == 'platform':
                platform_yaml = line[1]
            if line[0] == 'log_source':
                log_source_yaml = line[1]
            if line[0] == 'event_code':
                event_code_yaml = line[1]
            if line[0] == 'event_version':
                event_version_yaml = line[1]
            if control == 'data dictionary':
                if line == (None,None,None,None,None):
                    continue
                dict = {'standard_name' : line[0],
                        'name' : line[1],
                        'type' : line[2],
                        'description' : line[3],
                        'sample_value' : line[4]}
                event_fields_yaml.append(dict)
            if control == 'references':
                if line == (None,None,None,None,None):
                    continue
                references_dict_yaml = {'text':line[0],'link':line[1]}
                references_yaml.append(references_dict_yaml)
            if control == 'tags':
                tags_yaml.append(line[0].rstrip())
        # Dictionary of data to create yaml file
        data_dict = {'title' : title_yaml,
                    'description' : description_yaml,
                    'platform' : platform_yaml,
                    'log_source' : log_source_yaml,
                    'event_code' : event_code_yaml,
                    'event_version' : event_version_yaml,
                    'event_fields' : event_fields_yaml,
                    'references' : references_yaml,
                    'tags' : tags_yaml}
        # Formatting sheet name
        sheet = sheet.replace(' ','_')
        # Creating yaml file
        with open(sheet + '.yaml', 'w') as file:
            yaml.dump(data_dict, file, sort_keys = False, width = float("inf"))